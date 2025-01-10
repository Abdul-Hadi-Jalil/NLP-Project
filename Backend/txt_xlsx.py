from os.path import exists
import openpyxl

def txt_to_xlsx(input_file, output_file, delimiter=' '):
    try:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active  # Get the active worksheet
        
        # Open the .txt file
        with open(input_file, 'r') as file:
            for row_idx, line in enumerate(file, start=1):
                # Split the line into values based on the delimiter
                values = line.strip().split(delimiter)
                
                # Write each value to the corresponding cell
                for col_idx, value in enumerate(values, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook as an Excel file
        workbook.save(output_file)
    
    except FileNotFoundError:
        print(f"Error: The file {input_file} does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")


def xlsx_to_txt(xlsx_file, txt_file, delimiter=' '):
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(xlsx_file)
        
        # Get the active worksheet (or specify another sheet if needed)
        sheet = workbook.active
        
        # Open the .txt file in write mode
        with open(txt_file, 'w') as outfile:
            # Iterate through the rows in the sheet
            for row in sheet.iter_rows(values_only=True):  # values_only=True extracts just the cell values
                # Join the row values with the specified delimiter
                row_data = delimiter.join([str(cell) if cell is not None else "" for cell in row])
                # Write the row data to the .txt file
                outfile.write(row_data + '\n')  # Write each row as a new line

        print(f"File {xlsx_file} successfully converted to {txt_file}")

    except FileNotFoundError:
        print(f"Error: The file {xlsx_file} does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")
                

def main():
    input_file = input("Enter the file name or path which you want to convert (input.txt): ")
    output_file = input("Enter the file name or path in which you want to convert (input.xlsx): ")

    if input_file.endswith('.txt') and (output_file.endswith('.xls') or output_file.endswith('.xlsx')):
        try:
            if not exists(input_file):
                print(f"The specified file {input_file} does not exist in your path")
                return
            
            txt_to_xlsx(input_file, output_file)
            print(f"Your file {input_file} is successfully converted to {output_file}")
        except Exception as e:
            print(f"Error: {e}")
    else:
        print("The given format is not supported for conversion")


if __name__=="__main__":
    main()