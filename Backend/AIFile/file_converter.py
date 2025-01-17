import csv
import openpyxl
import os

class FileConverter:
    def convert(self, input_file, output_file):
        # Determine the input and output formats from file extensions
        input_format = input_file.split('.')[-1].lower()
        output_format = output_file.split('.')[-1].lower()

        # Ensure input and output formats are not the same
        if input_format == output_format:
            print("Input and output formats are the same. No conversion needed.")
            return

        # Call the corresponding method based on the formats
        if input_format == 'csv' and output_format == 'xlsx':
            self.csv_to_xlsx(input_file, output_file)
        elif input_format == 'xlsx' and output_format == 'csv':
            self.xlsx_to_csv(input_file, output_file)
        elif input_format == 'txt' and output_format == 'csv':
            self.txt_to_csv(input_file, output_file)
        elif input_format == 'csv' and output_format == 'txt':
            self.csv_to_txt(input_file, output_file)
        elif input_format == 'txt' and output_format == 'xlsx':
            self.txt_to_xlsx(input_file, output_file)
        elif input_format == 'xlsx' and output_format == 'txt':
            self.xlsx_to_txt(input_file, output_file)
        else:
            print(f"Conversion from '{input_format}' to '{output_format}' is not supported.")

    def csv_to_xlsx(self, input_file, output_file):
        try:
            with open(input_file, 'r') as csv_file:
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                for row_idx, line in enumerate(csv.reader(csv_file), start=1):
                    for col_idx, value in enumerate(line, start=1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
                workbook.save(output_file)
            print(f"Successfully converted '{input_file}' to '{output_file}'.")
        except Exception as e:
            print(f"Error: {e}")

    def xlsx_to_csv(self, input_file, output_file):
        try:
            workbook = openpyxl.load_workbook(input_file)
            worksheet = workbook.active
            with open(output_file, 'w', newline='') as csv_file:
                writer = csv.writer(csv_file)
                for line in worksheet.iter_rows(values_only=True):
                    writer.writerow(line)
            print(f"Successfully converted '{input_file}' to '{output_file}'.")
        except Exception as e:
            print(f"Error: {e}")

    def txt_to_csv(self, input_file, output_file):
        try:
            with open(input_file, 'r') as infile, open(output_file, 'w', newline='') as outfile:
                writer = csv.writer(outfile)
                for line in infile:
                    line = line.strip().split()  # Split line into a list of values
                    writer.writerow(line)  # Write the list as a row in the CSV
            print(f"Successfully converted '{input_file}' to '{output_file}'.")
        except Exception as e:
            print(f"Error: {e}")

    def csv_to_txt(self, input_file, output_file):
        try:
            with open(input_file, 'r') as infile, open(output_file, 'w', newline='') as outfile:
                reader = csv.reader(infile)
                for line in reader:
                    line = ' '.join(line)
                    outfile.write(line + '\n')
            print(f"Successfully converted '{input_file}' to '{output_file}'.")
        except Exception as e:
            print(f"Error: {e}")

    def txt_to_xlsx(self, input_file, output_file, delimiter=' '):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            with open(input_file, 'r') as file:
                for row_idx, line in enumerate(file, start=1):
                    values = line.strip().split(delimiter)
                    for col_idx, value in enumerate(values, start=1):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
            workbook.save(output_file)
            print(f"Successfully converted '{input_file}' to '{output_file}'.")
        except Exception as e:
            print(f"Error: {e}")

    def xlsx_to_txt(self, input_file, output_file, delimiter=' '):
        try:
            workbook = openpyxl.load_workbook(input_file)
            sheet = workbook.active
            with open(output_file, 'w') as outfile:
                for row in sheet.iter_rows(values_only=True):
                    row_data = delimiter.join([str(cell) if cell is not None else "" for cell in row])
                    outfile.write(row_data + '\n')
            print(f"Successfully converted '{input_file}' to '{output_file}'.")
        except Exception as e:
            print(f"Error: {e}")


# Main function to run the conversion
def main():
    input_file = input("Enter the file name or path which you want to convert: ")
    output_file = input("Enter the file name or path to which you want to convert: ")

    # Create the FileConverter instance
    converter = FileConverter()
    
    # Perform the conversion
    converter.convert(input_file, output_file)

if __name__ == "__main__":
    main()
