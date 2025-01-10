import csv
import openpyxl

def csv_to_xlsx(csv_file: str, xlsx_file: str = None):
    try:
        with open(csv_file, 'r') as input_file:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            for row_idx, line in enumerate(csv.reader(input_file), start=1):
                for col_idx, word in enumerate(line, start=1):
                    worksheet.cell(row=row_idx, column=col_idx, value=word)
            
            workbook.save(xlsx_file if xlsx_file else csv_file.replace('.csv', '.xlsx'))
            print(f"Successfully converted '{csv_file}' to '{xlsx_file}'.")

    except Exception as e:
        print(f"Error: {e}")


def xlsx_to_csv(xlsx_file: str, csv_file: str = None):
    try:
        workbook = openpyxl.load_workbook(xlsx_file)
        worksheet = workbook.active

        with open(csv_file if csv_file else xlsx_file.replace('.xlsx', '.csv'), 'w', newline='') as output_file:
            writer = csv.writer(output_file)
            for line in worksheet.iter_rows(values_only=True):
                writer.writerow(line)

            print(f"Successfully converted '{xlsx_file}' to '{csv_file}'.")
    except Exception as e:
        print(f"Error: {e}")


def main():
    # Provide file paths for testing
    csv_input = "sample.csv"  # Replace with the actual path to your CSV file
    xlsx_output = "sample_output.xlsx"  # Optional, replace with the desired XLSX output path

    xlsx_input = "sample.xlsx"  # Replace with the actual path to your XLSX file
    csv_output = "sample_output.csv"  # Optional, replace with the desired CSV output path

    # Test CSV to XLSX conversion
    print("Testing CSV to XLSX conversion:")
    csv_to_xlsx(csv_input, xlsx_output)

    # Test XLSX to CSV conversion
    print("\nTesting XLSX to CSV conversion:")
    xlsx_to_csv(xlsx_input, csv_output)


if __name__ == "__main__":
    main()
